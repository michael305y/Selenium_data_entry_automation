### using selenium to automate the web
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

from selenium.webdriver.support.ui import Select


###
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
###

# Create a Service object
service = webdriver.chrome.service.Service(ChromeDriverManager().install())

# Pass in the Service object to the webdriver.Chrome method
driver = webdriver.Chrome(service=service)

# driver.get("http://localhost:8000")

# driver.get("http://localhost:8000/show_KCPE_collection_form")  # FOR KCPE


driver.get("http://localhost:8000/show_KEPSEA_collection_form")   # FOR KEPSEA

driver.get("http://localhost:8000/show_KCSE_collection_form")   # FOR KCSE


# ====================

# # # Add a delay to allow the page to load
# # time.sleep(3)

# ===== search box =====
# search_box = driver.find_element(By.XPATH, '//*[@id="schoolCode"]')  

# search_box.send_keys('20409')

# ========================== GOOGLE SIGN ==========================================
# # sign in
# sign_in_btn = driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/div/div[2]/a/strong')
# sign_in_btn.click()

# # Log in with google button
# Login_btn = driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/a[1]')
# Login_btn.click()

# time.sleep(2)

# # continue button sign in via google
# continue_btn = driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/form/button')
# continue_btn.click()

# time.sleep(2)

# # # sign in textbox via google
# continue_btn = driver.find_element(By.XPATH, '//*[@id="identifierId"]')
# continue_btn.send_keys('kylesungunyo@gmail.com')

# # next button
# next_btn = driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/div[2]/div/c-wiz/div/div[2]/div/div[2]/div/div[1]/div/div/button/span')
# next_btn.click()
# ===============================END OF  GOOGLE SIGN IN ====================================================


# ====================== ENTERING DATA INTO THE SITE ======
# OPERATIONS TO ENTER DATA
# # school code textbox
# school_code = driver.find_element(By.XPATH, '//*[@id="id_school_code"]')
# school_code.send_keys("j")

# # school name textbox
# school_name = driver.find_element(By.XPATH, '//*[@id="id_school_name"]')
# school_name.send_keys("j")

# # school entry textbox
# school_entry = driver.find_element(By.XPATH, '//*[@id="id_entry"]')
# school_entry.send_keys("j")

# # school collection point combobox
# school_collection_point = driver.find_element(By.XPATH, '//*[@id="id_collection_point"]')
# school_collection_point.send_keys("j")

# # school route textbox
# school_route = driver.find_element(By.XPATH, '//*[@id="id_route"]')
# school_route.send_keys("j")

# # SUBMIT BTN
# submit_btn = driver.find_element(By.XPATH, '/html/body/div[1]/div/form/button')
# # submit_btn.

# # SUCCESS MESSAGE BTN
# success_message_btn = driver.find_element(By.XPATH, '/html/body/div[2]/div/div[4]/div/button')
# # success_message_btn.

# # INSERT COLLECTION BTN
# inssert_collection_btn = driver.find_element(By.XPATH, '/html/body/div[1]/a')
# # inssert_collection_btn.


# EXCEL OPERATION
workbook = openpyxl.load_workbook('kcpe_collection points.xlsx')

# Access the worksheet with the links
links_worksheet = workbook['Sheet1']

# /======== PRINTING ON THE TERMINAL FOR TESTING PURPSOSES ===========================================/
# Iterate over the rows in the sheet, starting from the second row
# for row in links_worksheet.iter_rows(min_row=2, max_row=links_worksheet.max_row, values_only=True):
#     # Check if all values in the row are None, indicating an empty row
#     if all(value is None for value in row):
#         continue  # Skip the empty row

#     print(row)
# /========END OF PRINTING ON THE TERMINAL FOR TESTING PURPSOSES ===========================================/

    
for row in links_worksheet.iter_rows(min_row=2, max_row=links_worksheet.max_row, values_only=True):
    SCHOOLDODE = row[0]
    SCHOOLNAME = row[1]
    ENTRY = row[2]
    COLLECTIONPOINT = row[3]
    ROUTE = row[4]
                                                 
    SCHOOLDODE = driver.find_element(By.XPATH, '//*[@id="id_school_code"]').send_keys(SCHOOLDODE)
                  
    time.sleep(1)

    SCHOOLNAME = driver.find_element(By.XPATH, '//*[@id="id_school_name"]').send_keys(SCHOOLNAME)

    ENTRY = driver.find_element(By.XPATH, '//*[@id="id_entry"]').send_keys(ENTRY)

    COLLECTIONPOINT = driver.find_element(By.XPATH, '//*[@id="id_collection_point"]').send_keys(COLLECTIONPOINT)

    ROUTE = driver.find_element(By.XPATH, '//*[@id="id_route"]').send_keys(ROUTE)

    time.sleep(1)  

    # # SUBMIT BTN
    submit_btn = driver.find_element(By.XPATH, '/html/body/div[1]/div/form/button')
    # submit_btn.click()  # incase submit button refuses to be clickable use option 2 with JS below
    driver.execute_script("arguments[0].click();", submit_btn) # using JS to click the submit button instead as option 2

    time.sleep(1)
    
    # SUCCESS MESSAGE BTN
    success_message_btn = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div/div[4]/div/button")))

    # success_message_btn = driver.find_element(By.XPATH, '/html/body/div[2]/div/div[4]/div/button')
    success_message_btn.click()

    time.sleep(1)

    # INSERT COLLECTION BTN
    insert_collection_btn = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/a")))
    
    # inssert_collection_btn = driver.find_element(By.XPATH, '/html/body/div[1]/a')
    insert_collection_btn.click()

    time.sleep(1)

 
while True:
    command = input("Enter Crtl + C to quit")
    # time.sleep(1)



   





